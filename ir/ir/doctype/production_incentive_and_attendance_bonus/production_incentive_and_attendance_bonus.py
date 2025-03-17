# Copyright (c) 2024, TEAMPROO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import (
    cstr, add_days, date_diff, getdate, format_date, flt, cint, 
    nowdate, today
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from six import BytesIO
from datetime import datetime

class ProductionIncentiveandAttendanceBonus(Document):
    pass

@frappe.whitelist()
def download():
    filename = 'Production Incentive and Attendance Bonus'
    build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or 'Sheet1'

    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50

    # Add headers and formatted date range
    ws.append(["Production Incentive and Attendance Bonus", " ", " "])

    args = frappe.local.form_dict
    from_date_str, to_date_str = args.from_date, args.to_date
    from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    
    formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")

    header3 = [f"For the Period: {formatted_from_date} to {formatted_to_date}"]
    ws.append(header3 + [""] * 3)
    ws.append([''])
    ws.append(['Employee', 'Production Incentive', 'Attendance Bonus'])

    # Append data
    for row in data:
        ws.append(row)

    # Apply styles to header cells
    align_center = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for header in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_center

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

    # Save to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def build_xlsx_response(filename):
    data = get_data(frappe.local.form_dict)
    xlsx_file = make_xlsx(data)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_data(args):
    data = []
    employees = get_employees(args.from_date)
    incen = 0
    bonus = 0
    for emp in employees:
        
        present_days = frappe.db.sql("""
            SELECT COUNT(status) AS present_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2 AND employee = %s
            AND status = 'Present'
        """, (args.from_date, args.to_date, emp.name), as_dict=True)[0].get("present_days", 0)
        
        paid_days = frappe.db.sql("""
            SELECT COUNT(status) AS paid_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2 AND employee = %s
            AND status = 'On Leave' AND leave_type != 'Leave Without Pay'
        """, (args.from_date, args.to_date, emp.name), as_dict=True)[0].get("paid_days", 0)
        
        leave_days = frappe.db.sql("""
            SELECT COUNT(status) AS leave_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2 AND employee = %s
            AND status = 'On Leave' AND leave_type = 'Leave Without Pay'
        """, (args.from_date, args.to_date, emp.name), as_dict=True)[0].get("leave_days", 0)
        
        half_day = frappe.db.sql("""
            SELECT SUM(CASE
                        WHEN leave_type != 'Leave Without Pay' THEN 1
                        WHEN leave_type = 'Leave Without Pay' THEN 0.5
                        ELSE 0
                    END) AS half_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2
            AND employee = %s
            AND status = 'Half Day'
        """, (args.from_date, args.to_date, emp.name), as_dict=True)[0].get("half_days", 0)

        
        attendance_dates = frappe.db.sql("""
            SELECT attendance_date
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2
            AND employee = %s
            AND (status = 'Present' OR status = 'On Leave')
        """, (args.from_date, args.to_date, emp.name), as_dict=True)


        holiday_list = frappe.db.get_value('Employee', {'name': emp.name}, 'holiday_list')
        holidays = frappe.db.sql("""
            SELECT holiday_date
            FROM `tabHoliday`
            WHERE parent = %s AND holiday_date BETWEEN %s AND %s
        """, (holiday_list, args.from_date, args.to_date), as_dict=True)

        comp_off_days = sum(1 for h in holidays if any(a.attendance_date == h.holiday_date for a in attendance_dates))
        comp_off_days = sum(1 for h in holidays if any(a.attendance_date == h.holiday_date for a in attendance_dates))
        total_present_days = present_days + len(holidays) - comp_off_days 
        payment_days = total_present_days + paid_days + (half_day or 0) 
        # Calculate incentive and attendance bonus
        total_days = date_diff(args.to_date, args.from_date) + 1
        incen += round((emp.custom_production_incentive)*(payment_days/total_days),2)
        if emp.custom_employee_category == "White Collar":
            attendance_bonus = 0
        else:
            attendance_bonus = 300 if total_present_days == total_days else 0
        bonus += attendance_bonus
        if emp.custom_production_incentive > 0 or attendance_bonus >0:
            # row = [emp.name,payment_days,total_days]
            row = [emp.name, round((emp.custom_production_incentive)*(payment_days/total_days),2) or 0, attendance_bonus]

            data.append(row)
    data.append(["Total",incen,bonus])
    return data

def get_employees(from_date):
    active_employees = frappe.db.sql("""
        SELECT * FROM `tabEmployee`
        WHERE custom_normal_employee = 1 AND status = 'Active'
    """, as_dict=True)
    
    left_employees = frappe.db.sql("""
        SELECT * FROM `tabEmployee`
        WHERE custom_normal_employee = 1 AND status = 'Left' 
        AND relieving_date >= %s
    """, from_date, as_dict=True)

    return active_employees + left_employees
