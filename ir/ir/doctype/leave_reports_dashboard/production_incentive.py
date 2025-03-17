from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from io import BytesIO
import frappe
from datetime import datetime
from frappe.utils import date_diff

@frappe.whitelist()
def download_production_incentive(from_date, to_date):
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
    filename = f'Production Incentive - {from_date_obj.strftime("%Y-%m")}'
    xlsx_file, total_incentive = make_xlsx(from_date, to_date)
    frappe.response['filename'] = f"{filename}.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Incentive"

    headers = {
        "main": ["Sl. No.", "Emp ID", "Employee Name", "CC", "Account No.", "Net Amount"],
    }

    set_headers(ws, headers, from_date)
    
    data, total_incentive = get_data({"from_date": from_date, "to_date": to_date})
    for row_idx, row in enumerate(data, start=12):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_border_font(cell)

    last_row = len(data) + 12
    add_total_row(ws, last_row, total_incentive)

    apply_table_border(ws, 12, 1, last_row, len(headers["main"]))
            
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file, total_incentive

def add_total_row(ws, last_row, total_incentive):
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
    label_cell = ws.cell(row=last_row, column=1, value="Total Amount")
    total_cell = ws.cell(row=last_row, column=6, value=total_incentive)
    
    footer1 = last_row + 5
    ws.merge_cells(start_row=footer1+1, start_column=1, end_row=footer1+1, end_column=2)
    prepared = ws.cell(row=footer1+1, column=1, value="Prepared By:")
    ws.merge_cells(start_row=footer1+1, start_column=3, end_row=footer1+1, end_column=4)
    checked = ws.cell(row=footer1+1, column=3, value="Checked By:")
    ws.merge_cells(start_row=footer1+1, start_column=5, end_row=footer1+1, end_column=6)
    approved = ws.cell(row=footer1+1, column=5, value="Approved By:")
    
    footer2 = footer1 + 2
    ws.merge_cells(start_row = footer2+1, start_column = 2, end_row = footer2+1, end_column = 5)
    company = ws.cell(row=footer2+1, column = 2, value="INDUSTRIAS DEL RECAMBIO INDIA PVT LTD")
    
    footer3 = footer2 + 4
    ws.merge_cells(start_row = footer3+1, start_column = 3, end_row = footer3+1, end_column = 4)
    sign = ws.cell(row=footer3+1, column = 3, value="Authorised Signatory:")
    
    
    bold_font = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell1 in [label_cell, total_cell]:
        cell1.font = bold_font
        cell1.alignment = alignment_center
        cell1.border = border
        
    for cell2 in [company]:
        cell2.font = bold_font
        cell2.alignment = alignment_center
        
    for cell3 in [prepared, approved, checked, sign]:
        cell3.font = bold_font
        cell3.alignment = alignment_left


def apply_cell_border_font(cell):
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    cell.font = bold_font
    cell.border = thin_border

def apply_table_border(ws, start_row, start_column, end_row, end_column):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            ws.cell(row=row, column=col).border = thin_border

def set_headers(ws, headers, from_date):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FABF8F", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = "INDUSTRIAS DEL RECAMBIO INDIA PVT LTD"
    ws['A1'].font = header_font
    ws['A1'].alignment = alignment_center
    apply_range_border(ws, 'A1:F1', thin_border)

    ws.merge_cells('A2:F2')
    ws['A2'] = "Bank Statement"
    ws['A2'].font = header_font
    ws['A2'].alignment = alignment_center
    apply_range_border(ws, 'A2:F2', thin_border)

    ws['A4'] = "To"
    ws['A4'].font = header_font
    ws['A4'].alignment = alignment_left

    ws.merge_cells('A5:B5')
    ws['A5'] = "Branch Manager"
    ws['A5'].font = header_font
    ws['A5'].alignment = alignment_left

    ws['A6'] = "Axis Bank"
    ws['A6'].font = header_font
    ws['A6'].alignment = alignment_left

    ws.merge_cells('A9:F9')
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
    ws['A9'] = (
        f"Please credit the following SB Accounts maintained with you by the amounts "
        f"mentioned against the account numbers. This is towards the production incentive "
        f"for the Month of {from_date_obj.strftime('%B %Y')}."
    )
    ws['A9'].font = header_font
    ws['A9'].alignment = alignment_center
    ws.row_dimensions[9].height = 30

    add_headers(ws, headers["main"], 11, header_font, header_fill, alignment_center, thin_border)

    column_widths = [10, 12, 30, 15, 20, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_headers(ws, headers, row_num, font, fill, alignment, border):
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        apply_cell_style(cell, font, fill, alignment, border)

def apply_cell_style(cell, font, fill, alignment, border):
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

def apply_range_border(ws, cell_range, border_style):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = border_style

def get_data(args):
    data = []
    employees = get_employees(args["from_date"])
    total_incentive = 0
    index = 1

    for emp in employees:
        emp_id = emp['name']
        emp_name = emp['employee_name']
        cc = emp['payroll_cost_center']
        account_no = emp['bank_ac_no']

        # Calculate present days
        present_days = frappe.db.sql(""" 
            SELECT COUNT(status) AS present_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2 AND employee = %s
            AND status = 'Present'
        """, (args["from_date"], args["to_date"], emp_id), as_dict=True)[0].get("present_days", 0)

        paid_days = frappe.db.sql(""" 
            SELECT COUNT(status) AS paid_days
            FROM `tabAttendance`
            WHERE attendance_date BETWEEN %s AND %s
            AND docstatus != 2 AND employee = %s
            AND status = 'On Leave' AND leave_type != 'Leave Without Pay'
        """, (args["from_date"], args["to_date"], emp_id), as_dict=True)[0].get("paid_days", 0)

        holidays = frappe.db.sql(""" 
            SELECT holiday_date
            FROM `tabHoliday`
            WHERE parent = %s AND holiday_date BETWEEN %s AND %s
        """, (
            frappe.db.get_value('Employee', {'name': emp_id}, 'holiday_list'),
            args["from_date"],
            args["to_date"],
        ), as_dict=True)

        comp_off_days = sum(1 for h in holidays if any(
            a.attendance_date == h.holiday_date for a in frappe.db.sql(""" 
                SELECT attendance_date
                FROM `tabAttendance`
                WHERE attendance_date BETWEEN %s AND %s
                AND docstatus != 2 AND employee = %s
                AND (status = 'Present' OR status = 'On Leave')
            """, (args["from_date"], args["to_date"], emp_id), as_dict=True)
        ))

        total_present_days = present_days + len(holidays) - comp_off_days
        payment_days = total_present_days + paid_days

        # Calculate Production Incentive
        total_days = date_diff(args["to_date"], args["from_date"]) + 1
        production_incentive = round(emp['custom_production_incentive'] * (payment_days / total_days))

        if production_incentive > 0:
            data.append([index, emp_id, emp_name, cc, account_no, production_incentive])
            total_incentive += production_incentive
            index += 1

    return data, total_incentive

def get_employees(from_date):
    active_employees = frappe.db.sql(""" 
        SELECT name, employee_name, payroll_cost_center, bank_ac_no, custom_production_incentive
        FROM `tabEmployee`
        WHERE custom_normal_employee = 1 AND status = 'Active'
    """, as_dict=True)
    
    left_employees = frappe.db.sql(""" 
        SELECT name, employee_name, payroll_cost_center, bank_ac_no, custom_production_incentive
        FROM `tabEmployee`
        WHERE custom_normal_employee = 1 AND status = 'Left' 
        AND relieving_date >= %s
    """, (from_date), as_dict=True)

    return active_employees + left_employees
