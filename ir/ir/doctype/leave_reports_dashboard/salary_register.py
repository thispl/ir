from frappe.model.document import Document
import frappe
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from six import BytesIO
from datetime import datetime

@frappe.whitelist()
def download_salary_register():
    filename = 'Salary Register.xlsx'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
    args = frappe.local.form_dict
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name if sheet_name else 'Sheet1'
    
    data = get_data(args)
    
    if not data:
        frappe.throw("No data available to generate the report.")
        
    # Header formatting
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=29)
    ws['A1'] = "INDUSTRIAS DEL RECAMBIO INDIA PVT LTD"
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws['A1'].font = Font(bold=True, size=14)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=29)
    from_date_obj = datetime.strptime(args.from_date, '%Y-%m-%d')
    ws['A2'] = f"Salary Sheet Report for the month of {from_date_obj.strftime('%B %Y')}."
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header row formatting for bold text
    header = ["SI. NO", "Emp ID", "Employee Name", "Employee Category","Cost Center", "Group", "Pay Days", 
             "Present Days", "BASIC+DA", "HRA", "Travel", "CCA", "Other Earn", "Other","Income Tax","Salary Arrear", "Spl Allowa", "OT1", "Total Earnings", 
             "PF", "ESI","Other/ Advance Deduction", "Professional Tax", "TDS", "Canteen","Labour Welfare Fund", "LIC", "Total Deductions", "Net Amount"]
    
    row_index = 5
    for col_index, header_cell in enumerate(header, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=header_cell)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)  # Apply bold font to the header cells
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))  # Apply border
    
    row_index += 1  # Move to the next row after header

    # Logic to track section headings (White, Blue, Grey Collars)
    white_collar_added = blue_collar_added = grey_collar_added = False
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    
    for row in data:
        if isinstance(row, str) and row == "White Collar" and not white_collar_added:
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=23)
            ws.cell(row=row_index, column=1).value = "White Collar"
            ws.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_index, column=1).font = Font(bold=True, size=12)
            white_collar_added = True
            row_index += 1
            
        elif isinstance(row, str) and row == "Blue Collar" and not blue_collar_added:
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=23)
            ws.cell(row=row_index, column=1).value = "Blue Collar"
            ws.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_index, column=1).font = Font(bold=True, size=12)
            blue_collar_added = True
            row_index += 1
            
        elif isinstance(row, str) and row == "Grey Collar" and not grey_collar_added:
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=23)
            ws.cell(row=row_index, column=1).value = "Grey Collar"
            ws.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_index, column=1).font = Font(bold=True, size=12)
            grey_collar_added = True
            row_index += 1
        
        elif isinstance(row, list):
            # Check if it's subtotal or grand total
            is_subtotal_or_grandtotal = row[0] in ["Sub Total", "Grand Total"]
            
            for col_index, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Do not apply borders for subtotal/grand total rows
                if is_subtotal_or_grandtotal:
                    cell.border = Border(left=Side(border_style=None), right=Side(border_style=None),
                                         top=Side(border_style=None), bottom=Side(border_style=None))
                else:
                    cell.border = border  # Apply borders for other rows
            row_index += 1
            
    last_row = len(data) + 5
    add_total_row(ws, last_row)

    # Save the final file as bytes and return
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def add_total_row(ws, last_row):
    footer1 = last_row + 5
    ws.merge_cells(start_row=footer1 + 1, start_column=5, end_row=footer1 + 1, end_column=8)
    prepared = ws.cell(row=footer1 + 1, column=5, value="Prepared By:")
    ws.merge_cells(start_row=footer1 + 1, start_column=9, end_row=footer1 + 1, end_column=12)
    checked = ws.cell(row=footer1 + 1, column=9, value="Checked By:")
    ws.merge_cells(start_row=footer1 + 1, start_column=13, end_row=footer1 + 1, end_column=16)
    approved = ws.cell(row=footer1 + 1, column=13, value="Verified By:")
    ws.merge_cells(start_row=footer1 + 1, start_column=17, end_row=footer1 + 1, end_column=22)
    company = ws.cell(row=footer1 + 1, column=17, value="For INDUSTRIAS DEL RECAMBIO INDIA PVT LTD")

    footer3 = footer1 + 2
    ws.merge_cells(start_row=footer3 + 1, start_column=17, end_row=footer3 + 1, end_column=19)
    sign = ws.cell(row=footer3 + 1, column=17, value="Authorised Signatory:")
    
    bold_font = Font(bold=True, size=12)
    bold_font1 = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell1 in [prepared, checked, approved, sign]:
        cell1.font = bold_font1
        cell1.alignment = alignment_center
        
    for cell in [company]:
        cell.font = bold_font
        cell.alignment = alignment_center

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(sheet_name=filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def get_data(args):
    from_date = args.get('from_date')
    to_date = args.get('to_date')

    if not from_date or not to_date:
        frappe.throw("From Date and To Date are required.")
    
    data = []  # Remove header from data initialization here
    
    category = ['White Collar', 'Blue Collar', 'Grey Collar']
    
    grand_total = []
    g_total_other_advance = g_total_income_tax = g_total_salary_arrear = g_total_pay_days = g_total_present_days = g_total_basic = g_total_hra = g_total_travelling = g_total_cca = g_total_pf = g_total_tds = g_total_canteen = g_total_lwf = g_total_lic = g_earn_total = g_dec_total = g_net_total = g_total_ot = g_total_esi = g_total_pt = g_total_other= g_total_earn= g_total_spe= 0
    
    for cat in category:
        salary_slips = frappe.db.sql(""" 
            SELECT * 
            FROM `tabSalary Slip` 
            WHERE start_date = %s AND end_date = %s 
            AND docstatus != 2 AND custom_employee_category = %s
        """, (from_date, to_date, cat), as_dict=True)
        
        total_other_advance = total_income_tax = total_salary_arrear = total_pay_days = total_present_days = total_basic = total_hra = total_travelling = total_cca = total_pf = total_tds = total_canteen = total_lwf =  total_lic = earn_total = dec_total = net_total = total_ot = total_esi = total_pt = total_other=total_earn=total_spe= 0
        
        index = 0
        sub_total = []
        data.append(cat)  # Only append category once
        
        for s in salary_slips:
            if not isinstance(s, dict):
                continue

            emp_data = frappe.get_list(
                "Employee",
                filters={"name": s['employee']},
                fields=["name", "employee_name", "payroll_cost_center", "custom_group", "custom_employee_category"],
                limit_page_length=1
            )
            
            emp = emp_data[0] if emp_data else frappe._dict({
                'name': None,
                'employee_name': None,
                'payroll_cost_center': None,
                'custom_group': None,
                'custom_employee_category': None
            })

            components = [
                "Basic", "House Rent Allowance", "Travelling Allowance", "CCA","Other Earn", "Other","Earning Income Tax","Salary Arrear", "Special Allowance",
                "OT", "Total Earning", "Deduction PF", "Deduction ESI","Other/ Advance Deduction", "Professional Tax", "Tax Deducted Source", "Canteen","Labour Welfare Fund", "LIC"
            ]

            salary_details = {}
            for comp in components:
                salary_amount = frappe.get_value("Salary Detail", {'parent': s['name'], 'salary_component': comp}, 'amount') or 0.0
                salary_details[comp] = salary_amount

            total_earnings = sum(salary_details.get(comp, 0.0) for comp in ["Basic", "House Rent Allowance", "Travelling Allowance", "CCA", "Other"])
         
            index += 1

            row = [
                index, 
                s['employee'], 
                s['employee_name'], 
                s['custom_employee_category'],
                emp['payroll_cost_center'],
                emp['custom_group'],
                s['payment_days'], 
                s.get('custom_present_days_holiday', 0), 
                salary_details.get("Basic", 0),
                salary_details.get("House Rent Allowance", 0), 
                salary_details.get("Travelling Allowance", 0),
                salary_details.get("CCA", 0),
                salary_details.get("Other Earn", 0),
                salary_details.get("Other", 0),
                salary_details.get("Earning Income Tax", 0),
                salary_details.get("Salary Arrear", 0),
                salary_details.get("Special Allowance", 0),
                salary_details.get("OT", 0),
                s['gross_pay'],
                salary_details.get("Deduction PF", 0), 
                salary_details.get("Deduction ESI", 0),
                salary_details.get("Other/ Advance Deduction",0),
                salary_details.get("Professional Tax", 0), 
                salary_details.get("Tax Deducted Source", 0),
                salary_details.get("Canteen", 0), 
                salary_details.get("Labour Welfare Fund", 0), 
                salary_details.get("LIC", 0), 
                s['total_deduction'], 
                s['net_pay']
            ]
            data.append(row)
            total_pay_days += s['payment_days']
            total_present_days += s.get('custom_present_days_holiday', 0)
            total_basic += salary_details.get("Basic", 0)
            total_hra += salary_details.get("House Rent Allowance", 0)
            total_travelling += salary_details.get("Travelling Allowance", 0)
            total_cca += salary_details.get("CCA", 0)
            total_earn+=salary_details.get("Other Earn", 0)
            total_other+=salary_details.get("Other", 0)
            total_spe+=salary_details.get("Special Allowance", 0)
            total_income_tax+=salary_details.get("Earning Income Tax", 0)
            total_salary_arrear+=salary_details.get("Salary Arrear", 0)
            total_ot += salary_details.get("OT", 0)
            earn_total += s['gross_pay']
            total_pf += salary_details.get("Deduction PF", 0)
            total_esi += salary_details.get("Deduction ESI", 0)
            total_other_advance += salary_details.get("Other/ Advance Deduction",0)
            total_pt += salary_details.get("Professional Tax", 0)
            total_tds += salary_details.get("Tax Deducted Source", 0)
            total_canteen += salary_details.get("Canteen", 0)
            total_lwf += salary_details.get("Labour Welfare Fund", 0)
            total_lic +=  salary_details.get("LIC", 0)
            dec_total += s['total_deduction']
            net_total += s['net_pay']
            
            g_total_pay_days += s['payment_days']
            g_total_present_days += s.get('custom_present_days_holiday', 0)
            g_total_basic += salary_details.get("Basic", 0)
            g_total_hra += salary_details.get("House Rent Allowance", 0)
            g_total_travelling += salary_details.get("Travelling Allowance", 0)
            g_total_cca += salary_details.get("CCA", 0)
            g_total_earn+=salary_details.get("Other Earn", 0)
            g_total_other+=salary_details.get("Other", 0)
            g_total_spe+=salary_details.get("Special Allowance", 0)
            g_total_income_tax+=salary_details.get("Earning Income Tax", 0)
            g_total_salary_arrear+=salary_details.get("Salary Arrear", 0)
            g_total_ot += salary_details.get("OT", 0)
            g_earn_total += s['gross_pay']
            g_total_pf += salary_details.get("Deduction PF", 0)
            g_total_esi += salary_details.get("Deduction ESI", 0)
            g_total_other_advance += salary_details.get("Other/ Advance Deduction",0)
            g_total_pt += salary_details.get("Professional Tax", 0)
            g_total_tds += salary_details.get("Tax Deducted Source", 0)
            g_total_canteen += salary_details.get("Canteen", 0)
            g_total_lwf += salary_details.get("Labour Welfare Fund", 0)
            g_total_lic +=  salary_details.get("LIC", 0)
            g_dec_total += s['total_deduction']
            g_net_total += s['net_pay']
            
        # sub_total += ["Sub Total", "", "", "", "", total_pay_days, total_present_days, total_basic, total_hra, total_travelling, total_cca,
        #                 "", "", "", total_ot, earn_total, total_pf, total_esi, total_pt, total_tds, total_canteen,total_lwf, total_lic, dec_total, net_total]
        sub_total += ["Sub Total", "", "", "", "","",total_pay_days, total_present_days, total_basic, total_hra, total_travelling, total_cca,
                      total_earn,total_other,total_income_tax,total_salary_arrear,total_spe,total_ot, earn_total, total_pf, total_esi,total_other_advance, total_pt, total_tds, total_canteen,total_lwf, total_lic, dec_total, net_total]
        data.append(sub_total)
            
    # grand_total += ["Grand Total", "", "", "", "", g_total_pay_days, g_total_present_days, g_total_basic, g_total_hra, g_total_travelling, g_total_cca,
    #                     "", "", "", g_total_ot, g_earn_total, g_total_pf, g_total_esi, g_total_pt, g_total_tds, g_total_canteen,g_total_lwf, g_total_lic, g_dec_total, g_net_total]
    grand_total += ["Grand Total", "", "", "", "","", g_total_pay_days, g_total_present_days, g_total_basic, g_total_hra, g_total_travelling, g_total_cca,
                        g_total_earn,g_total_other,g_total_income_tax,g_total_salary_arrear,g_total_spe, g_total_ot, g_earn_total, g_total_pf, g_total_esi,g_total_other_advance, g_total_pt, g_total_tds, g_total_canteen,g_total_lwf, g_total_lic, g_dec_total, g_net_total]
    data.append(grand_total)
        
    return data
