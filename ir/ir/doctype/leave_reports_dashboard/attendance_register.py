from __future__ import unicode_literals
from functools import total_ordering
from itertools import count
import frappe
from frappe import permissions
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from math import floor
from frappe import msgprint, _
from calendar import month, monthrange
from datetime import date, timedelta, datetime,time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, GradientFill, PatternFill
from six import BytesIO
from frappe.utils.background_jobs import enqueue

@frappe.whitelist()
def download_attendance_register(from_date, to_date, employee):
    filename = 'Attendance Register'
    args = {"from_date": from_date, "to_date": to_date, "employee": employee}
    # build_xlsx_response(filename)
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response', filename=filename, args=args)

def make_xlsx(args,sheet_name=None):
    # args = frappe.local.form_dict
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    from_date = args["from_date"]
    to_date = args["to_date"]
    employee = args["employee"]

    data = get_data(from_date, to_date, employee)
    head = hearders(from_date, to_date, employee)

    ws.append(head)

    for row in data:
        ws.append(row)
    
    bold_font = Font(bold=True)
    align_center = Alignment(horizontal='center',vertical='center')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'))
    
    for cell in ws[1]:
        cell.font = bold_font
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align_center
            cell.border = border
 
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

# def build_xlsx_response(filename):
#     xlsx_file = make_xlsx(sheet_name= "Attendance Register")
#     frappe.response['filename'] = filename + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'
    
def build_xlsx_response(filename,args):
	xlsx_file = make_xlsx(args,sheet_name= None)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Leave Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Leave Reports Dashboard',None,'attach',attached_file.file_url)

status_map = {
    'Permission Request' :'PR',
    'On Duty Application':'OD',
    'Half Day':'HD',
    "Absent": "A",
    "Half Day": "HD",
    "Holiday": "HH",
    "Weekly Off": "WW",
    "Present": "P",
    "None" : "",
    "Leave Without Pay": "LOP",
    "Casual Leave": "CL",
    "Earned Leave": "EL",
    "Sick Leave": "SL",
    "Emergency -1": 'EML-1',
    "Emergency -2": 'EML-2',
    "Paternal Leave": 'PL',
    "Marriage Leave":'ML',
    "Medical Leave" : 'MDL',
    "Paternity Leave":'PTL',
    "Education Leave":'EL',
    "Maternity Leave":'MTL',
    "Covid -19": "COV-19",
    "Privilege Leave": "PVL",
    "Compensatory Off": "C-OFF",
    "BEREAVEMENT LEAVE":'BL'
}

def hearders(from_date, to_date, employee):
    headers = []
    headers += [
        "Employee ID", "Employee Name", "Department", "DOJ", "Status"
    ]
    dates = get_dates(from_date, to_date, employee)
    for date in dates:
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')

        day = date.strftime('%d')
        month = date.strftime('%b')
        headers.append(f"{day}/{month}")

    headers.append("Payment Days")
    headers.append("Present")
    headers.append('Paid Leave')
    headers.append('Paid Holiday')
    headers.append('Weekoff')
    headers.append('COFF')
    headers.append("Absent")
    headers.append('LOP')
    headers.append('Night Shift')
    headers.append('Total OT Hours')
    headers.append('Total Permission Hours')
    
    return headers


def get_data(from_date, to_date, employee):
    data = []
    employees = get_employees(from_date, to_date, employee)
    dates = get_dates(from_date, to_date, employee)
    for emp in employees:
        row1 = [emp.name,emp.employee_name,emp.department,emp.date_of_joining,""]
        row2 = ["","","","","In Time"]
        row3 = ["","","","","Out Time"]
        row4 = ["","","","","Shift"]
        row5 = ["","","","","TWH"]
        row6 = ["","","","","EH"]
        row7 = ["","","","","OT"]
        row8 = ["","","","","Permission Hour"]
        row9 = ["","","","","Late"]
        row10 = ["","","","","Early Out"]
        total_present = 0
        total_half_day = 0
        total_absent = 0
        total_holiday = 0
        total_weekoff = 0
        weekoff = 0
        total_ot = 0
        total_od = 0
        total_permission = 0
        total_lop = 0
        total_paid_leave = 0
        total_combo_off = 0
        c_shift = 0
        total_late = timedelta(0,0,0)
        total_late_deduct = timedelta(0,0)
        ww = 0
        twh = 0
        ot = 0
        pay=0
        per = 0
        for date in dates:
            od = frappe.db.get_value("On Duty Application",{'from_date':date,'employee':emp.name,'docstatus':('!=','2')},['from_date_session'])
            ot_ = frappe.db.get_value("Over Time Request",{'ot_date':date,'employee':emp.name,'workflow_state':('Not In',('Rejected','Cancelled'))},['ot_hour'])
            att = frappe.db.get_value("Attendance",{'attendance_date':date,'employee':emp.name,'docstatus':('!=','2')},['status','in_time','out_time','shift','custom_total_working_hours','leave_type','employee','attendance_date','name','custom_ot_hours','custom_permission_hour','late_entry',"leave_application","custom_on_duty_application",'custom_attendance_permission','custom_late_entry_time','custom_extra_hours_total','custom_early_out_time']) or ''
            if att:
                if att[17]:
                  row10.append(att[17]) 
                else:
                    row10.append("-")   
                if att[11]:
                    row9.append(att[11])
                else:
                    row9.append("-")
                if att[10]:
                    # per = int(att[10])
                    row8.append(att[10])
                else:
                    row8.append("-")
                if att[9]:
                    row7.append(att[9])
                    # ot+=att[9]
                    ftr = [3600,60,1]
                    if ot_:
                        hr = sum(a * b for a, b in zip(ftr, map(float, str(ot_).split(':'))))
                        ot_hr = round(hr/3600,1)
                        ot += ot_hr
                else:
                    row7.append("-")
                if att[16]:
                    row6.append(att[16])
                else:
                    row6.append("-")
                status = status_map.get(att[0], "")    
                if att[0] == 'Present':
                    # pay +=1
                    
                    # else:
                    # 	row1.append('P')
                    hh = check_holiday(date,emp.name)
                    if hh:
                        if hh == 'WW':
                            row1.append(hh)
                            total_weekoff +=1
                            if att[0] == "Present":
                                weekoff -=1
                            elif att[0] == "Half Day":
                                weekoff -=0.5
                        if hh == 'HH':
                            row1.append(hh)
                            total_holiday +=1  
                            if att[0] == "Present":
                                weekoff -=1
                            elif att[0] == "Half Day":
                                weekoff -=0.5 
                        # row1.append(hh)   
                    else:
                        total_present +=1 
                        if att[13] is not None:
                            od = frappe.db.get_value("On Duty Application",{'docstatus':('!=','2'),"name":att[13]},['from_date_session'])
                            if od == 'Second Half':
                                row1.append('P/OD')
                            if od == 'First Half':
                                row1.append('OD/P')
                            if od  == 'Full Day':
                                row1.append('OD')
                        elif att[14]:
                            pr = frappe.db.get_value("Permission Request",{'docstatus':('!=','2'),"name":att[14],"employee":emp.name,"attendance_date":date},['session'])
                            if pr == 'Second Half':
                                row1.append('P/PR')
                            if pr == 'First Half':
                                row1.append('PR/P')
                        else:
                            row1.append(status)
                        # total_present = total_present + 1 

                elif att[0] == 'Half Day':
                    hh = check_holiday(date,emp.name)
                    if hh:
                        if hh == 'WW':
                            
                            row1.append(hh)
                            total_weekoff += 1
                        elif hh == 'HH':
                            row1.append(hh)
                            total_holiday += 1
                    elif[12]:
                        leave_applications = frappe.db.sql("""
                            SELECT leave_type 
                            FROM `tabLeave Application` 
                            WHERE from_date=%s AND employee=%s AND status='Approved'
                        """, (date, emp.name), as_dict=True)

                        if len(leave_applications) > 1:
                            leave_types = []
                            leave_type_map = {
                                "Earned Leave": "EL",
                                "Compensatory Off": "COFF",
                                "Casual Leave": "CL",
                                "Sick Leave": "SL",
                                "Leave Without Pay": "LOP"
                            }
                            
                            for leave in leave_applications:
                                leave_type_code = leave_type_map.get(leave['leave_type'], leave['leave_type'])
                                leave_types.append(leave_type_code)
                                if leave['leave_type'] == "Leave Without Pay":
                                    # total_paid_leave += 0.5
                                    total_lop += 0.5
                                else:
                                    total_paid_leave += 0.5
                            
                            lstatus = "/".join(leave_types)
                            row1.append(lstatus)

                        elif att[5] and att[1] and att[2]:
                            shift_et = frappe.db.get_value("Shift Type",{'name':att[3]},['start_time'])
                            time1 = datetime.strptime(str(att[1]),'%Y-%m-%d %H:%M:%S').time()
                            time2 = datetime.strptime(str(shift_et), '%H:%M:%S').time()
                            datetime1 = datetime.combine(datetime.min, time1)
                            datetime2 = datetime.combine(datetime.min, time2)
                            timedelta_seconds = (datetime1 - datetime2).total_seconds()
                            diff_timedelta = timedelta(seconds=timedelta_seconds)
                            if att[4] >=timedelta(hours=4):
                                if att[15] >= timedelta(hours=3):
                                    
                                    if att[5] == "Earned Leave": 
                                        row1.append('EL/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_paid_leave = total_paid_leave + 0.5	
                                    elif att[5] == "Compensatory Off":
                                        row1.append('COFF/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_paid_leave = total_paid_leave + 0.5	
                                    elif att[5] == "Casual Leave":
                                        row1.append('CL/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_paid_leave = total_paid_leave + 0.5	
                                    elif att[5] == "Sick Leave":
                                        row1.append('SL/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_paid_leave = total_paid_leave + 0.5	
                                    elif att[5] == "Leave Without Pay":
                                        row1.append('LOP/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                    else:
                                        row1.append('L/P')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                    
                                else:
                                    if att[5] == "Earned Leave": 
                                        row1.append('P/EL')
                                        total_half_day = total_half_day + 0.5
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5		
                                    elif att[5] == "Compensatory Off":
                                        row1.append('P/COFF')
                                        total_half_day = total_half_day + 0.5
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5		
                                    elif att[5] == "Casual Leave":
                                        row1.append('P/CL')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_paid_leave = total_paid_leave + 0.5		
                                    elif att[5] == "Sick Leave":
                                        row1.append('P/SL')
                                        total_half_day = total_half_day + 0.5
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5	
                                    elif att[5] == "Leave Without Pay":
                                        row1.append('P/LOP')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                    else:
                                        row1.append('P/L')
                                        total_half_day = total_half_day + 0.5
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                    
                            else:
                                if att[14]:
                                    if att[5] == "Earned Leave": 
                                        row1.append('P/EL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5	
                                    elif att[5] == "Compensatory Off":
                                        row1.append('P/COFF')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5	
                                    elif att[5] == "Casual Leave":
                                        row1.append('P/CL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5	
                                    elif att[5] == "Sick Leave":
                                        row1.append('P/SL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_present=total_present+ 0.5	
                                    elif att[5] == "Leave Without Pay":
                                        row1.append('P/LOP')
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                    else:
                                        row1.append('A/L')
                                        total_present=total_present+ 0.5
                                        total_lop += 0.5
                                else:
                                    if att[5] == "Earned Leave": 
                                        row1.append('A/EL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_absent=total_absent+ 0.5	
                                    elif att[5] == "Compensatory Off":
                                        row1.append('A/COFF')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_absent=total_absent+ 0.5	
                                    elif att[5] == "Casual Leave":
                                        row1.append('A/CL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_absent=total_absent+ 0.5	
                                    elif att[5] == "Sick Leave":
                                        row1.append('A/SL')
                                        total_paid_leave = total_paid_leave + 0.5
                                        total_absent=total_absent+ 0.5	
                                    elif att[5] == "Leave Without Pay":
                                        row1.append('A/LOP')
                                        total_absent=total_absent+ 0.5
                                        total_lop += 0.5
                                    else:
                                        row1.append('A/L')
                                        total_absent=total_absent+ 0.5
                                        total_lop += 0.5
                               
                        elif att[5] and not att[2] or not att[1]:
                            if att[5] == "Earned Leave": 
                                row1.append('A/EL')
                                total_paid_leave = total_paid_leave + 0.5
                                total_absent=total_absent+ 0.5	
                            elif att[5] == "Compensatory Off":
                                row1.append('A/COFF')
                                total_paid_leave = total_paid_leave + 0.5
                                total_absent=total_absent+ 0.5	
                            elif att[5] == "Casual Leave":
                                row1.append('A/CL')
                                total_paid_leave = total_paid_leave + 0.5
                                total_absent=total_absent+ 0.5	
                            elif att[5] == "Sick Leave":
                                row1.append('A/SL')
                                total_paid_leave = total_paid_leave + 0.5
                                total_absent=total_absent+ 0.5	
                            elif att[5] == "Leave Without Pay":
                                row1.append('A/LOP')
                                total_lop = total_lop + 0.5
                                total_absent=total_absent+ 0.5
                            else:
                                row1.append('A/L')
                                total_lop = total_lop + 0.5
                                total_absent=total_absent+ 0.5
                        elif att[13] and not att[2] or not att[1]:
                            od = frappe.db.get_value("On Duty Application",{'docstatus':('!=','2'),"name":att[13]},['from_date_session'])
                            if od == 'Second Half':
                                row1.append('A/OD')
                            if od == 'First Half':
                                row1.append('OD/A')
                            if od  == 'Full Day':
                                row1.append('OD')
                            total_present +=0.5
                        elif att[13] and att[2] and  att[1]:
                            
                            od = frappe.db.get_value("On Duty Application",{'docstatus':('!=','2'),"name":att[13]},['from_date_session'])
                            time = frappe.db.get_value("On Duty Application",{'docstatus':('!=','2'),"name":att[13]},['od_time'])
                            if time > 4 :
                                if od == 'Second Half':
                                
                                    if att[4] >=timedelta(hours=4):
                                        
                                            row1.append('P/OD')
                                    else:
                                        row1.append('A/OD')
                            
                                        
                                if od == 'First Half':
                                    if att[4] >=timedelta(hours=4):
                                        row1.append('OD/P')
                                    else:
                                        row1.append('OD/A')
                                if od  == 'Full Day':
                                    row1.append('OD')
                                total_present = total_present + 0.5
                            else:
                                if od == 'Second Half':
                                
                                    if att[4] >=timedelta(hours=4):
                                        
                                            row1.append('P/A')
                                    else:
                                        row1.append('A')
                            
                                        
                                if od == 'First Half':
                                    if att[4] >=timedelta(hours=4):
                                        row1.append('OD/A')
                                    else:
                                        row1.append('A')
                                if od  == 'Full Day':
                                    row1.append('OD/A')
                                total_present = total_present + 0.5
                        elif att[14]:
                            pr = frappe.db.get_value("Permission Request",{'docstatus':('!=','2'),"name":att[14],"employee":emp.name,"attendance_date":date},['session'])
                    
                            if pr == 'Second Half':
                                row1.append('A/PR')
                            if pr == 'First Half':
                                row1.append('PR/A')
                            total_present = total_present + 0.5
                        else:
                            if att[11] == 1:
                                row1.append('A/P')
                            else:
                                row1.append('P/A')
                            total_half_day = total_half_day + 0.5
                            total_absent=total_absent+ 0.5
                            total_present = total_present + 0.5
                            
                    else:
                       total_present = total_present + 0.5
                       total_absent=total_absent+ 0.5   
                elif att[0] == 'Absent':
                    hh = check_holiday(date,emp.name)
                    
                    if hh:
                        # pay +=1
                        if hh == 'WW':
                            
                            row1.append(hh)
                            total_weekoff += 1
                        elif hh == 'HH':
                            row1.append(hh)
                            total_holiday += 1
                        # row1.append(hh)

                    elif att[13]:
                        od = frappe.db.get_value("On Duty Application",{'docstatus':('!=','2'),"name":att[13]},['from_date_session'])
                        if od == 'Second Half':
                            if att[4] >=timedelta(hours=4):
                                row1.append('P/OD')
                            else:
                                row1.append('A/OD')
                        if od == 'First Half':
                            if att[4] >=timedelta(hours=4):
                                row1.append('OD/P')
                            else:
                                row1.append('OD/A')
                        if od  == 'Full Day':
                            row1.append('OD')
                        total_present = total_present + 0.5
                        total_absent = total_absent + 0.5
                    else:    
                        row1.append(status)
                        total_absent = total_absent + 1                         
                
                elif att[0] == "On Leave":
                    if att[5] == "Earned Leave": 
                        row1.append('EL')
                        total_paid_leave = total_paid_leave + 1	
                    elif att[5] == "Compensatory Off":
                        row1.append('COFF')
                        total_paid_leave = total_paid_leave + 1	
                    elif att[5] == "Casual Leave":
                        row1.append('CL')
                        total_paid_leave = total_paid_leave + 1	
                    elif att[5] == "Sick Leave":
                        row1.append('SL')
                        total_paid_leave = total_paid_leave + 1
                    elif att[5] == "Leave Without Pay":
                        row1.append('LOP')
                        total_lop +=1
                    else:
                        row1.append('-')
                elif att[5]:
                    hh = check_holiday(date,emp.name)
                    if hh:
                        # pay+=1
                        if hh == 'WW':
                            
                            total_weekoff += 1
                            if att[0] == "Present":
                                weekoff -=1
                            elif att[0] == "Half Day":
                                weekoff -=0.5
                        elif hh == 'HH':
                            total_holiday += 1
                            if att[0] == "Present":
                                weekoff -=1
                            elif att[0] == "Half Day":
                                weekoff -=0.5
                        row1.append(hh)
                    else:    
                        status = status_map.get(att[5], "")
                        if status != 'LOP':
                            if status == 'C-OFF':
                                total_combo_off += 1
                                total_paid_leave += 1
                            else:
                                # pay+=1
                                total_paid_leave += 1
                        else:                        
                            total_lop += 1
                        row1.append(status)
                else:
                    row1.append('-')
                if att[1] is not None :
                    if att[1] != '0':
                        row2.append(att[1].strftime('%H:%M'))
                    else:
                        row2.append('-')

                else:
                    row2.append('-')
                if att[2] is not None :
                    if att[2] != '0':
                        row3.append(att[2].strftime('%H:%M'))
                    else:
                        row3.append('-')

                else:
                    row3.append('-')
                
                if att[3]:
                    if att[0] == 'Absent':
                        row4.append(att[3] or '')
                    else:
                        row4.append(att[3])    
                else:
                    row4.append('-')

                if att[3] == 'C':
                    c_shift += 1       

                if att[4] is not None:
                    hh = att[4].seconds//3600
                    mm = (att[4].seconds//60)%60
                    twh = str(hh) + ":" + str(mm)
                    row5.append(twh) 
                else:
                    row5.append('-')    
            
            else:
                # row7.append("-")
                hh = check_holiday(date,emp.name)
                if hh:
                    # pay+=1
                    if hh == 'WW': 
                        
                        total_weekoff += 1
                    
                    elif hh == 'HH':
                        total_holiday += 1
            
                    row1.append(hh)
                    holiday = frappe.db.get_value('Holiday Attendance',{'employee':emp.name,'attendance_date':date,'docstatus':('!=','2')},['in_time','out_time','shift','late_hours','late_deduct','total_wh','ot_hrs'])
                    if holiday:
                        if holiday[0]:
                            row2.append(holiday[0].strftime('%H:%M'))
                        else:
                            row2.append('-')  
                        if holiday[1]:
                            row3.append(holiday[1].strftime('%H:%M'))
                        else:
                            row3.append('-')  
                        if holiday[2]:
                            row4.append(holiday[2])   
                        else:
                            row4.append('-')   
                        if holiday[3]:
                            row5.append(holiday[3]) 
                        else:
                            row5.append('-')            
                    else:
                        row2.append('-')
                        row3.append('-') 
                        row4.append('-') 
                        row5.append('-')
                        row6.append('-')
                        row7.append('-')
                        row8.append('-')
                        row9.append('-')
                        row10.append('-')
                else:
                    row1.append('-')
                    row2.append('-')
                    row3.append('-')
                    row4.append('-')
                    row5.append('-')
                    row6.append('-')
                    row7.append('-')
                    row8.append('-')
                    row9.append('-')
                    row10.append('-')
        total_payment_days = total_present+total_weekoff+total_holiday+total_paid_leave
        permission_hours = frappe.db.sql("""select sum(hours) as sum from `tabPermission Request` where attendance_date between '%s' and '%s' and employee = '%s' and docstatus = '1' """%(from_date,to_date,emp.name),as_dict=True)[0].sum or 0
        ot_request = frappe.db.sql("""select sum(TIME_TO_SEC(ot_hour)) / 3600 as total_hours from `tabOver Time Request` where ot_date between '%s' and '%s' and employee = '%s' and docstatus = '1' """%(from_date,to_date,emp.name),as_dict=True)
        per = ot_request[0].total_hours or 0
        row1.extend([total_present+total_weekoff+total_holiday+total_paid_leave,total_present,total_paid_leave,total_holiday,total_weekoff,total_combo_off,total_absent,total_lop,c_shift or 0.0,ot or 0.0,per or 0])
        row2.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row3.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row4.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row5.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row6.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row7.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row8.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row9.extend(['-','-','-','-','-','-','-','-','-','-','-'])
        row10.extend(['-','-','-','-','-','-','-','-','-','-','-'])


        data.append(row1)
        data.append(row2)
        data.append(row3)
        data.append(row4)
        data.append(row5)
        data.append(row6)
        data.append(row7)
        data.append(row8)
        data.append(row9)
        data.append(row10)
    return data

def get_dates(from_date, to_date, employee):
    no_of_days = date_diff(add_days(to_date, 1), from_date)
    dates = [add_days(from_date, i) for i in range(0, no_of_days)]
    
    return dates

def get_employees(from_date, to_date, employee):
    conditions = ''
    left_employees = []
    if employee:
        conditions += "and employee = '%s' " % (employee)
    employees = frappe.db.sql("""select name, employee_name, department,employment_type,date_of_joining from `tabEmployee` where employment_type !='Agency' AND  status = 'Active' %s """ % (conditions), as_dict=True)
    left_employees = frappe.db.sql("""select name, employee_name, department,employment_type,date_of_joining from `tabEmployee` where employment_type !='Agency' AND status = 'Left' and relieving_date >= '%s' %s """ %(from_date, conditions),as_dict=True)
    employees.extend(left_employees)
    return employees
  
@frappe.whitelist()
def get_to_date(from_date):
    day = from_date[-2:]
    if int(day) > 21:
        d = add_days(get_last_day(from_date),21)
        return d
    if int(day) <= 21:
        d = add_days(get_first_day(from_date),21)
        return d

def check_holiday(date,emp):
    holiday_list = frappe.db.get_value('Company','Industrias Del Recambio India Private Limited','default_holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    doj= frappe.db.get_value("Employee",{'name':emp},"date_of_joining")
    last= frappe.db.get_value("Employee",{'name':emp},"relieving_date")
    if holiday:
        if last:
            if doj < holiday[0].holiday_date < last:
                if holiday[0].weekly_off == 1:
                    return "WW"
                else:
                    return "HH"
        else:
            if doj < holiday[0].holiday_date:
                if holiday[0].weekly_off == 1:
                    return "WW"
                else:
                    return "HH"

    

    
    


