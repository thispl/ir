from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import frappe
import math
from datetime import datetime

@frappe.whitelist()
def download_esi_monthly(from_date, to_date):
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
    filename = f'ESI Monthly - {from_date_obj.strftime("%Y-%m")}'
    xlsx_file = make_xlsx(from_date, to_date)
    frappe.response['filename'] = f"{filename}.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "ESI Report"
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')

    headers = {
        "first_head": ["INDUSTRIAS DEL RECAMBIO INDIA PVT LTD", "Employer ESI No.: 51001058050000699"],
        "month": [f"ESI report for the month of {from_date_obj.strftime('%B %Y')}"],
        "main": ["S.No", "EMP", "CC", "ESI No", "Name of Member", "Days Worked ", "ESI Earnings", "ESI Contribution for employee ", "ESI Contribution for Employer"],
        "sub_sub": ["", "", "", "", "", "", "", "", ""]
    }

    set_headers(ws, headers)
    data_rows = gather_employee_data(from_date, to_date)
    write_data(ws, data_rows)
    
    # Add totals row
    add_totals(ws, len(data_rows) + 6)  # Length of data_rows + starting row index (6)
    
    adjust_column_widths(ws)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def set_headers(ws, headers):
    # Define styles once
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="f8f8ff", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Merge the first two rows into a single row
    ws.merge_cells('A1:I2')
    ws['A1'] = headers['first_head'][0] + " - " + headers['first_head'][1]
    apply_cell_style(ws['A1'], header_font, header_fill, alignment_center, thin_border)

    # Add month header on the third row (unchanged)
    ws['A3'] = headers['month'][0]
    ws.merge_cells('A3:I3')
    apply_cell_style(ws['A3'], header_font, header_fill, alignment_center, thin_border)

    # Main headers on the fourth row
    add_headers(ws, headers['main'], 4, header_font, header_fill, alignment_center, thin_border)

    # Sub headers on the fifth row
    add_headers(ws, headers['sub_sub'], 5, header_font, header_fill, alignment_center, thin_border)

def apply_cell_style(cell, font, fill, alignment, border):
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

def add_headers(ws, headers, row_num, font, fill, alignment, border, merge_ranges=None):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        apply_cell_style(cell, font, fill, alignment, border)
    if merge_ranges:
        for start_row, start_col, end_row, end_col in merge_ranges:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

def gather_employee_data(from_date, to_date):
    salary_slips = frappe.db.get_all(
        'Salary Slip',
        filters={'docstatus': ('!=', '2'), 'start_date': from_date, 'end_date': to_date},
        fields=['name', 'employee_name', 'employee', 'payment_days', 'gross_pay']
    )
    employee_data = {
        emp['name']: emp for emp in frappe.db.get_all(
            'Employee',
            filters={'status': 'Active'},
            fields=['name', 'payroll_cost_center', 'custom_esi_no', 'custom_is_applicable_for_esi']
        )
    }

    data_rows = []
    index = 1  # Ensure numbering starts from 1
    for slip in salary_slips:
        employee_info = employee_data.get(slip['employee'])
        if not employee_info or not employee_info.get('custom_is_applicable_for_esi', False):
            continue  # Skip employees who are not applicable for ESI

        earnings = frappe.db.get_all(
            'Salary Detail',
            filters={'parent': slip['name']},
            fields=['amount', 'salary_component']
        )
        esi_no = employee_info['custom_esi_no']
        pay_days = slip['payment_days']
        gross = slip['gross_pay']
        esi_deduction = next((e['amount'] for e in earnings if e['salary_component'] == 'Deduction ESI'), 0)
        employer_contribution = gross * 3.25 / 100

        # Append the row with a properly ordered index
        data_rows.append([
            index,  # Ordered index
            slip['employee'],
            employee_info.get('payroll_cost_center', ''),
            esi_no,
            slip['employee_name'],
            pay_days,
            gross,
            round(esi_deduction, 2),
            math.ceil(employer_contribution)
        ])
        index += 1  # Increment the index for the next row

    return data_rows


def write_data(ws, data):
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_num, row in enumerate(data, start=6):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = alignment_center
            cell.border = thin_border

def add_totals(ws, start_row):
    # Calculate totals for the numeric columns (index 6 to 10)
    total_employer_esi = sum(ws.cell(row=row, column=7).value or 0 for row in range(6, start_row))
    total_employee_contribution = sum(ws.cell(row=row, column=8).value or 0 for row in range(6, start_row))
    total_esi_earnings = sum(ws.cell(row=row, column=9).value or 0 for row in range(6, start_row))
    overall_total = total_employee_contribution + total_esi_earnings

    # Write totals row (one row below the last data row)
    totals_row = start_row + 1
    ws.cell(row=totals_row, column=1, value="Total")  # Label cell

    # Write totals in the appropriate columns
    ws.cell(row=totals_row, column=7, value=round(total_employer_esi,2))
    ws.cell(row=totals_row, column=8, value=round(total_employee_contribution,2))
    ws.cell(row=totals_row, column=9, value=round(total_esi_earnings,2))
    

    # Apply formatting to the totals row (bold font and center alignment)
    for col in range(1, 10):  # Adjust range to match total columns
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    account_no_row = totals_row + 1
    ws.cell(row=account_no_row, column=5, value="Employee Contribution : ")
    ws.cell(row=account_no_row, column=6, value=round(total_employee_contribution,2))

    # Formatting for "Account No: 01" row
    account_cell = ws.cell(row=account_no_row, column=6)
    account_cell.font = Font(bold=True)
    account_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add "EDLI Wages:" in the row below Account No: 01
    edli_wages_row = account_no_row + 1
    ws.cell(row=edli_wages_row, column=5, value="Employer Contribution : ")
    ws.cell(row=edli_wages_row, column=6, value=round(total_esi_earnings,2))

    # Formatting for "EDLI Wages:" row
    edli_cell = ws.cell(row=edli_wages_row, column=6)
    edli_cell.font = Font(bold=True)
    edli_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    # Add "EDLI Wages:" in the row below Account No: 01
    overall = edli_wages_row + 1
    ws.cell(row=overall, column=5, value="TOTAL : ")
    ws.cell(row=overall, column=6, value=round(overall_total,2))

    # Formatting for "EDLI Wages:" row
    edli_cell = ws.cell(row=overall, column=6)
    edli_cell.font = Font(bold=True)
    edli_cell.alignment = Alignment(horizontal='center', vertical='center')



def adjust_column_widths(ws):
    for col_num in range(1, 10):
        if col_num == 1:  # S.No column
            ws.column_dimensions[get_column_letter(col_num)].width = 5  # Set a fixed width for S.No
        else:
            max_length = max((len(str(ws.cell(row=row_num, column=col_num).value)) if ws.cell(row=row_num, column=col_num).value else 0) for row_num in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 3
