from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import frappe
import math

@frappe.whitelist()
def download_pf_monthly(from_date, to_date):
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
    filename = f'PF Monthly - {from_date_obj.strftime("%Y-%m")}'
    xlsx_file = make_xlsx(from_date, to_date)
    frappe.response['filename'] = f"{filename}.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "PF Report"
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')

    headers = {
        "first_head": ["INDUSTRIAS DEL RECAMBIO INDIA PVT LTD", "Employer PF No:TN/AMB/34196"],
        "month": [f"PF Report For The Month Of {from_date_obj.strftime('%B %Y')}"],
        "main": ["S.No", "EMP", "UAN", "Name of Member", "CC", "Employer PF Contribution", "Employee Contribution", "", "Employer Contribution", "Employer Contribution"],
        "sub_sub": ["", "", "", "", "", "", "PF Earnings", "Contribution EPF", "EPF Difference", "Pension Fund"]
    }

    set_headers(ws, headers)
    data_rows = gather_employee_data(from_date, to_date)
    write_data(ws, data_rows)
    
    # Add totals row
    add_totals(ws, len(data_rows) + 6)  # Length of data_rows + starting row index (6)
    
    adjust_column_widths(ws)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def set_headers(ws, headers):
    # Define styles once
    header_font = Font(bold=True, color="000000",size=14)
    header_fill = PatternFill(start_color="f8f8ff", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # thin_border_1= Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

    # Merge the first two rows into a single row
    ws.merge_cells('A1:J2')
    ws['A1'] = headers['first_head'][0] + " - " + headers['first_head'][1]
    apply_cell_style(ws['A1'], header_font, header_fill, alignment_center, thin_border)

    # Add month header on the third row (unchanged)
    ws['A3'] = headers['month'][0]
    ws.merge_cells('A3:J3')
    apply_cell_style(ws['A3'], header_font, header_fill, alignment_center, thin_border)

    # Main headers on the fourth row
    add_headers(ws, headers['main'], 4, header_font, header_fill, alignment_center, thin_border, merge_ranges=[(4, 7, 4, 8), (4, 9, 4, 10)])

    # Sub headers on the fifth row
    add_headers(ws, headers['sub_sub'], 5, header_font, header_fill, alignment_center, thin_border)

def apply_cell_style(cell, font, fill, alignment, border):
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

def add_headers(ws, headers, row_num, font, fill, alignment, border, merge_ranges=None):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        apply_cell_style(cell, font, fill, alignment, border)
    if merge_ranges:
        for start_row, start_col, end_row, end_col in merge_ranges:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

def gather_employee_data(from_date, to_date):
    salary_slips = frappe.db.get_all('Salary Slip', filters={'docstatus': ('!=', '2'), 'custom_agency': '', 'start_date': from_date, 'end_date': to_date}, fields=['name', 'employee_name', 'employee', 'custom_agency'])
    employee_ids = [slip['employee'] for slip in salary_slips]
    employee_data = {emp['name']: emp for emp in frappe.db.get_all('Employee', filters={'name': ('in', employee_ids)}, fields=['name', 'uan_no', 'payroll_cost_center', 'custom_basic', 'custom_not_applicable_for_pension'])}

    data_rows = []
    for index, slip in enumerate(salary_slips, start=1):
        employee_info = employee_data.get(slip['employee'])
        if not employee_info:
            continue  # Skip inactive or non-existent employees

        earnings = frappe.db.get_all('Salary Detail', filters={'parent': slip['name']}, fields=['amount', 'salary_component'])
        uan_no = employee_info['uan_no']
        custom_basic_salary = employee_info['custom_basic']
        custom_not_applicable_for_pension = employee_info['custom_not_applicable_for_pension']
        actual_basic_salary = min(custom_basic_salary, 15000)  # Cap at 15,000

        pf_deduction = next((e['amount'] for e in earnings if e['salary_component'] == 'Deduction PF'), 0)
        eared_basic = next((e['amount'] for e in earnings if e['salary_component'] == 'Basic'), 0)
        epf_contribution = actual_basic_salary  # Employer PF Contribution
        
        
        if eared_basic >= 15000:
            basic = 15000
        else:
            basic = eared_basic
        eps_and_epf = basic * 8.33 / 100 if not custom_not_applicable_for_pension else 0
        if eps_and_epf == 0:
            epf_contribution_1 = basic * 12 / 100
        else:
            epf_contribution_1 = basic * 3.67 / 100
        
        data_rows.append([index,
                          slip['employee'],
                          uan_no,
                          slip['employee_name'],
                          employee_info.get('payroll_cost_center', ''),
                          round(basic,2),
                          round(eared_basic,2),
                          round(pf_deduction,2),
                          round(epf_contribution_1),
                          round(eps_and_epf)])
    return data_rows

def write_data(ws, data):
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_num, row in enumerate(data, start=6):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = alignment_center
            cell.border = thin_border

def add_totals(ws, start_row):
    # Calculate totals for the numeric columns (index 6 to 10)
    total_employer_pf = sum(ws.cell(row=row, column=6).value or 0 for row in range(6, start_row))
    total_employee_contribution = sum(ws.cell(row=row, column=7).value or 0 for row in range(6, start_row))
    total_pf_earnings = sum(ws.cell(row=row, column=8).value or 0 for row in range(6, start_row))
    total_contribution_epf = sum(ws.cell(row=row, column=9).value or 0 for row in range(6, start_row))
    total_pension_fund = sum(ws.cell(row=row, column=10).value or 0 for row in range(6, start_row))
    total_employees = sum(1 for row in range(6, start_row) if ws.cell(row=row, column=1).value)  # Count non-empty employee rows
    total_combined = round(total_pf_earnings + total_contribution_epf,2)
    account_no_02_value = total_employer_pf * 0.5 / 100
    account_no_21_value = total_employer_pf * 0.5 / 100
    account_no_22_value = 0.00
    overall_total = round((total_combined + account_no_02_value + total_pension_fund + account_no_21_value),2)

    # Write totals row (one row below the last data row)
    totals_row = start_row + 1
    ws.cell(row=totals_row, column=1, value="Grand Total")  # Label cell
    ws.merge_cells(start_row=totals_row, start_column=2, end_row=totals_row, end_column=5)
    # ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=10)

    # Write totals in the appropriate columns
    ws.cell(row=totals_row, column=6, value=round(total_employer_pf,2))
    ws.cell(row=totals_row, column=7, value=round(total_employee_contribution,2))
    ws.cell(row=totals_row, column=8, value=round(total_pf_earnings,2))
    ws.cell(row=totals_row, column=9, value=round(total_contribution_epf,2))
    ws.cell(row=totals_row, column=10, value=round(total_pension_fund,2))

    # Apply formatting to the totals row (bold font and center alignment)
    for col in range(1, 12):  # Adjust range to match total columns
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    

    # Add "Account No: 01" in the row immediately below the totals row
    account_no_row = totals_row + 1
    ws.cell(row=account_no_row, column=9, value="Account No: 01 =")
    ws.cell(row=account_no_row, column=10, value=round(total_combined,2))
    ws.merge_cells(start_row=account_no_row, start_column=1, end_row=account_no_row, end_column=8)

    # Formatting for "Account No: 01" row
    account_cell = ws.cell(row=account_no_row, column=6)
    account_cell.font = Font(bold=True)
    account_cell.alignment = Alignment(horizontal='center', vertical='center')


    # Add "Account No: 02" in the row below EDLI Wages
    account_no_02_row = account_no_row + 1
    ws.cell(row=account_no_02_row, column=9, value="Account No: 02 = ")

    # Calculate and insert value for Account No: 02 based on 0.5% of EDLI Wages
    account_no_02_value = total_employer_pf * 0.50000 / 100
    ws.cell(row=account_no_02_row, column=10, value=round(account_no_02_value,2))
    ws.merge_cells(start_row=account_no_02_row, start_column=1, end_row=account_no_02_row, end_column=8)
    
        # Add "EDLI Wages:" in the row below Account No: 01
    edli_wages_row = account_no_02_row + 1
    ws.cell(row=edli_wages_row, column=1, value="EDLI Wages:")
    ws.cell(row=edli_wages_row, column=2, value=round(total_employer_pf,2))
    ws.merge_cells(start_row=edli_wages_row, start_column=3, end_row=edli_wages_row, end_column=8)
    

    # Formatting for "EDLI Wages:" row
    edli_cell = ws.cell(row=edli_wages_row, column=1)
    edli_cell.font = Font(bold=True)
    edli_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wages_row = edli_wages_row + 1
    ws.cell(row=wages_row, column=1, value="Pension Wages:")
    ws.cell(row=wages_row, column=2, value=0.00)
    ws.merge_cells(start_row=wages_row, start_column=3, end_row=wages_row, end_column=8)
    

    # Formatting for "EDLI Wages:" row
    wages_row_cell = ws.cell(row=wages_row, column=1)
    wages_row_cell.font = Font(bold=True)
    wages_row_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatting for "Account No: 02" row
    account_no_02_cell = ws.cell(row=account_no_02_row, column=6)
    account_no_02_cell.font = Font(bold=True)
    account_no_02_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add "Account No: 10" in the row below Account No: 02
    account_no_10_row = account_no_02_row + 1
    ws.cell(row=account_no_10_row, column=9, value="Account No: 10 =")
    ws.cell(row=account_no_10_row, column=10, value=round(total_pension_fund,2))
    # ws.merge_cells(start_row=account_no_row, start_column=1, end_row=account_no_row, end_column=8)

    # Formatting for "Account No: 10" row
    account_no_10_cell = ws.cell(row=account_no_10_row, column=6)
    account_no_10_cell.font = Font(bold=True)
    account_no_10_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add "Account No: 21" in the row below Account No: 10
    account_no_21_row = account_no_10_row + 1
    ws.cell(row=account_no_21_row, column=9, value="Account No: 21 = ")

    # Calculate and insert value for Account No: 21 based on 0.5% of EDLI Wages (you can change calculation as per your needs)
    account_no_21_value = total_employer_pf * 0.50000 / 100
    ws.cell(row=account_no_21_row, column=10, value=round(account_no_21_value,2))
    # ws.merge_cells(start_row=account_no_row, start_column=1, end_row=account_no_row, end_column=8)

    # Formatting for "Account No: 21" row
    account_no_21_cell = ws.cell(row=account_no_21_row, column=6)
    account_no_21_cell.font = Font(bold=True)
    account_no_21_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    # Add "Account No: 22"
    account_no_22_row = account_no_21_row + 1
    ws.cell(row=account_no_22_row, column=9, value="Account No: 22 =")
    ws.cell(row=account_no_22_row, column=10, value=account_no_22_value)
    ws.merge_cells(start_row=account_no_22_row, start_column=1, end_row=account_no_22_row, end_column=8)

    # Formatting for "Account No: 21" row
    account_no_22_cell = ws.cell(row=account_no_22_row, column=6)
    account_no_22_cell.font = Font(bold=True)
    account_no_22_cell.alignment = Alignment(horizontal='center', vertical='center')
    

    # Add "Total No. of Employees in the Month" in the row below Account No: 21
    overall_total_row = account_no_22_row + 1
    ws.cell(row=overall_total_row, column=1, value="Total No. of Employees in the Month:")  # Label cell for employee count
    
    # ws.merge_cells(start_row=overall_total_row, start_column=1, end_row=overall_total_row, end_column=10)

    # Write the total number of employees
    ws.cell(row=overall_total_row, column=2, value=total_employees)
    # ws.merge_cells(start_row=overall_total_row, start_column=3, end_row=overall_total_row, end_column=8)

    # Apply formatting to the overall total row (bold font and center alignment)
    for col in range(1, 12):  # Adjust range to match total columns
        cell = ws.cell(row=overall_total_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add "Total No. of Employees in the Month" in the row below Account No: 21
    overall_total_row = account_no_22_row + 1
    ws.cell(row=overall_total_row, column=8, value="TOTAL:")  # Label cell for employee count

    # Write the total number of employees
    ws.cell(row=overall_total_row, column=9, value=overall_total)

    # Apply formatting to the overall total row (bold font and center alignment)
    for col in range(1, 12):  # Adjust range to match total columns
        cell = ws.cell(row=overall_total_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

   

def adjust_column_widths(ws):
    for col_num in range(1, 11):
        if col_num == 1:  # S.No column
            ws.column_dimensions[get_column_letter(col_num)].width = 5  # Set a fixed width for S.No
        else:
            max_length = max((len(str(ws.cell(row=row_num, column=col_num).value)) if ws.cell(row=row_num, column=col_num).value else 0) for row_num in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 3
