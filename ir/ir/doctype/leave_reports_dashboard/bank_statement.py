from __future__ import unicode_literals
import frappe
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from six import BytesIO
from frappe.utils import cstr, getdate, format_date, add_days
import datetime
from datetime import date, timedelta, datetime, time

@frappe.whitelist()
def download_hr_to_accounts():
    filename = 'Bank_Statement.xlsx'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
    args = frappe.local.form_dict
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    data, total_incentive = get_data(args)  # Get data and total_incentive
    
    # Add data to sheet
    for row in data:
        ws.append(row)

    # Merging cells for headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)

    # Get the last row for total row
    last_row = len(data) + 1
    # Merge Total row's first 4 columns (Total row)
    # ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
    # label_cell = ws.cell(row=last_row, column=1, value="Total Amount")
    # total_cell = ws.cell(row=last_row, column=5, value=total_incentive)  # Ensure total_incentive is defined
    
    # Add the text "Total has been worked" to the merged cell in the total row
    # total_worked_cell = ws.cell(row=last_row, column=1, value="Total has been worked")
    
    # Apply the same styling to the total row
    footer1 = last_row + 5
    ws.merge_cells(start_row=footer1+1, start_column=1, end_row=footer1+1, end_column=2)
    prepared = ws.cell(row=footer1+1, column=1, value="Prepared By:")
    ws.merge_cells(start_row=footer1+1, start_column=3, end_row=footer1+1, end_column=4)
    checked = ws.cell(row=footer1+1, column=3, value="Checked By:")
    ws.merge_cells(start_row=footer1+1, start_column=5, end_row=footer1+1, end_column=6)
    approved = ws.cell(row=footer1+1, column=5, value="Approved By:")
    
    footer2 = footer1 + 2
    ws.merge_cells(start_row=footer2+1, start_column=2, end_row=footer2+1, end_column=5)
    company = ws.cell(row=footer2+1, column=2, value="INDUSTRIAS DEL RECAMBIO INDIA PVT LTD")
    
    footer3 = footer2 + 4
    ws.merge_cells(start_row=footer3+1, start_column=3, end_row=footer3+1, end_column=4)
    sign = ws.cell(row=footer3+1, column=3, value="Authorised Signatory:")
    
    
    bold_font = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell1 in []:
        cell1.font = bold_font
        cell1.alignment = alignment_center
        cell1.border = border
        
    for cell2 in [company]:
        cell2.font = bold_font
        cell2.alignment = alignment_center
        
    for cell3 in [prepared, approved, checked, sign]:
        cell3.font = bold_font
        cell3.alignment = alignment_left


    # Apply alignment and borders
    header_font = Font(bold=True, color="000000")
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_text = Alignment(wrap_text=True)

    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Style headers and total row
    for rows in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5):
        for cell in rows:
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = border
    for rows in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=5):
        for cell in rows:
            cell.font = Font(bold=True)
            cell.border = border
    for rows in ws.iter_rows(min_row=6, max_row=len(data), min_col=1, max_col=5):
        for cell in rows:
            cell.border = border

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    ws['A5'].font = header_font
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=5)
    ws['A5'].font = header_font
    # Merge and center row 6, apply wrap text
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=5)
    ws['A6'].alignment = wrap_text
    
    ws['A6'].alignment = alignment_left
    ws['A6'].alignment = alignment_center
    
    ws['A6'].font = Font(bold=True)
    ws['A6'].border = border
    ws.row_dimensions[6].height = 30

    # Set zoom scale and column widths
    ws.sheet_view.zoomScale = 100 
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30

    # Save file to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def get_data(args):
    # Format month and category
    month = datetime.strptime(str(args.from_date), '%Y-%m-%d')
    month_str = month.strftime('%B %Y')
    employee_category = args.employee_category or ' '
    bank = args.bank or ' '

    data = [
        ["Industrias Del Recambio India Private Limited"],
        ["Bank Statement - " + month_str],
        ["To"],
        ["Branch Manager"],
        ["Bank", "Axis Bank", "", "", ""],
        [
            f"Please credit the following SB Accounts maintained with you by the amounts "
            f"mentioned against the account numbers. This is towards the Bank Statement "
            f"for the Month of {month_str}."
        ],
        ["S No", "Employee", "Employee Name", "Account Number", "Net Amount"]
    ]
    
    employees = get_employees(args)
    total_incentive = 0  # Initialize the total_incentive variable
    for i, employee in enumerate(employees, start=1):
        total_incentive += employee.get("rounded_total", 0)
        employeecode = employee.get("employee")
        data.append([i, employee.get("employee"), employee.get("employee_name"), employee.get("bank_account_no"), employee.get("rounded_total")])
    
    # Add total row
    data.append(["Total Amount", "", "", "", total_incentive])  # Total row (first 4 columns merged)
    
    return data, total_incentive

@frappe.whitelist()
def get_employees(args):
    conditions = []
    if args.from_date:
        conditions.append("start_date = '%s'" % args.from_date)
    if args.to_date:
        conditions.append("end_date = '%s'" % args.to_date)
    if args.bank:
        conditions.append("bank_name = '%s'" % args.bank)
    if args.employee_category:
        conditions.append("custom_employee_category = '%s'" % args.employee_category)

    where_clause = " AND ".join(conditions)
    sql_query = """SELECT * FROM `tabSalary Slip` WHERE {0} and custom_employee_category != ' ' """.format(where_clause)
    
    try:
        employees = frappe.db.sql(sql_query, as_dict=True)
    except Exception as e:
        employees = []

    return employees