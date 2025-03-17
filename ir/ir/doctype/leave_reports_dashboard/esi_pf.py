from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import frappe

@frappe.whitelist()
def download_esi(from_date,to_date):
    filename = 'PF_Report'
    build_xlsx_response(filename,from_date,to_date)

def build_xlsx_response(filename,from_date,to_date):
    xlsx_file = make_xlsx(from_date, to_date)
    frappe.response['filename'] = f"{filename}.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "PF Report"

    # Define headers
    headers = [
        "S.No", "UAN", "Member Name", "Gross Wages", "EPF Wages",
        "EPS Wages", "EDLI Wages", "EPF Contribution remitted",
        "EPS Contribution remitted", "EPF and EPS Diff remitted",
        "Non Contribution Payday"
    ]

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")  # Updated color
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Fetch all active salary slips with related employee UAN
    salary_slips = frappe.db.get_all(
        'Salary Slip',
        filters={'docstatus': ('!=', '2'), 'custom_agency': '','start_date':from_date,'end_date':to_date},  # Assuming docstatus=0 means draft (active)
        fields=['name', 'employee_name', 'gross_pay', 'employee', 'leave_without_pay'],
    )

    # Fetch active employees' UANs
    employee_uans = frappe.db.get_all(
        'Employee',
        filters={
            'name': ['in', [slip.employee for slip in salary_slips]],
            'status': 'Active'  # Ensure only active employees are included
        },
        fields=['name', 'uan_no', 'custom_not_applicable_for_pension'],
    )

    # Create a dictionary for UAN lookup
    uan_dict = {emp['name']: emp for emp in employee_uans}

    # Write data to worksheet
    for index, slip in enumerate(salary_slips, start=1):
        # Fetch earnings for the current Salary Slip
        earnings = frappe.db.get_all(
            'Salary Detail',
            filters={'parent': slip.name},
            fields=['amount', 'salary_component'],
        )

        # Get employee data for UAN and custom_not_applicable_for_pension
        employee_data = uan_dict.get(slip.employee)
        uan_no = employee_data['uan_no'] if employee_data else None
        custom_not_applicable_for_pension = employee_data['custom_not_applicable_for_pension'] if employee_data else 0

        # Get Basic Salary and PF Deduction amounts
        basic_salary = next((earning['amount'] for earning in earnings if earning['salary_component'] == 'Basic'), 0)
        pf_deduction = next((earning['amount'] for earning in earnings if earning['salary_component'] == 'Deduction PF'), 0)
        if basic_salary >= 15000:
            basic = 15000
        else:
            basic = basic_salary
        eps_and_epf = basic * 8.33 / 100 if not custom_not_applicable_for_pension else 0
        # if eps_and_epf == 0:
        #     pf_deduction = basic * 12 / 100
        # else:
        #     pf_deduction = basic* 3.67 / 100
        # Cap Basic Salary at 15,000
        capped_basic_salary = basic
        
        # Calculate EPS Wages based on custom_not_applicable_for_pension
        eps_wages = 0 if custom_not_applicable_for_pension == 1 else capped_basic_salary

        eps_contribution_remitted = capped_basic_salary * 0.0833
        # if custom_not_applicable_for_pension == 0:
        #     eps_and_epf = capped_basic_salary * 3.67 / 100
        # else:
        #     eps_and_epf = capped_basic_salary * 12 / 100

        ws.append([
            index,
            uan_no,
            slip.employee_name,
            round(slip.gross_pay, 0),
            round(capped_basic_salary, 0),
            round(eps_wages, 0),  
            round(capped_basic_salary, 0),  
            round(pf_deduction, 0),
            round(eps_contribution_remitted, 0),
            round(eps_and_epf, 0),
            slip.leave_without_pay,
        ])

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    # Add border for the entire table
    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=len(salary_slips) + 1):
        for cell in row:
            cell.border = thin_border

    # Save the workbook to a BytesIO object for file handling
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)  # Move the cursor to the beginning of the file

    return xlsx_file
