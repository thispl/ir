from io import BytesIO
import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

@frappe.whitelist()
def download(from_date, to_date, employee):
    filename = 'Leave_Report'
    build_xlsx_response(filename, from_date, to_date, employee)

def make_xlsx(filename, from_date, to_date, employee):
  
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    # Main headers row
    headers = ["S.No", "Employee", "Name", "Designation", "Department", "DOJ", 
               "Opening Balance", "", "", "", 
               "Availed Leaves", "", "", "", 
               "Closing Balance", "", "", ""]
    
    # Sub-header row for each leave type
    sub_headers = ["", "", "", "", "", "",
                   "CL", "SL", "EL", "LOP",
                   "CL", "SL", "EL", "LOP",
                   "CL", "SL", "EL", "LOP"]

    # Add the headers to the worksheet
    ws.append(headers)
    ws.append(sub_headers)

    # Merging cells for the main headers
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=10)  # "Opening Balance"
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=14) # "Availed Leaves"
    ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=18) # "Closing Balance"

    # Apply formatting for the headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_num, sub_header in enumerate(sub_headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define border style for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Fetch and populate the leave data
    leave_data = get_leave_data(from_date, to_date, employee)
    row_num = 3
    for idx, row_data in enumerate(leave_data, 1):
        ws.append([
            idx,  # S.No
            row_data["employee"],
            row_data["employee_name"],
            row_data["designation"],
            row_data["department"],
            row_data["doj"],
            row_data["opening_cl"],
            row_data["opening_sl"],
            row_data["opening_el"],
            row_data["opening_lop"],
            row_data["availed_cl"],
            row_data["availed_sl"],
            row_data["availed_el"],
            row_data["availed_lop"],
            row_data["closing_cl"],
            row_data["closing_sl"],
            row_data["closing_el"],
            row_data["closing_lop"],
        ])

        # Apply border to all cells in the row
        for col_num in range(1, 19):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border

        row_num += 1

    # Save the workbook into a BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename, from_date, to_date, employee):
    xlsx_file = make_xlsx(filename, from_date, to_date, employee)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_leave_data(from_date, to_date, employee):
    
    employees = frappe.db.get_all('Employee', fields=['name', 'employee_name', 'designation', 'department', 'date_of_joining as doj'], 
                                  filters={'status': 'Active',"custom_normal_employee":1})

    leave_types = ["Casual Leave", "Sick Leave", "Earned Leave", "Leave Without Pay"]

    leave_data = []
    
    for employee in employees:
        
        # Data for each employee
        row_data = {
            "employee": employee['name'],
            "employee_name": employee['employee_name'],
            "designation": employee['designation'],
            "department": employee['department'],
            "doj": employee['doj'],
            "opening_cl": 0,
            "opening_sl": 0,
            "opening_el": 0,
            "opening_lop": 0,
            "availed_cl": 0,
            "availed_sl": 0,
            "availed_el": 0,
            "availed_lop": 0,
            "closing_cl": 0,
            "closing_sl": 0,
            "closing_el": 0,
            "closing_lop": 0
        }

        # Fetch data for each leave type
        for leave_type in leave_types:
            
            opening_balance = get_opening_balance(employee['name'], leave_type, from_date)
            availed_leaves = get_availed_leaves(employee['name'], leave_type, from_date, to_date)
            closing_balance = opening_balance - availed_leaves
            
            if leave_type == "Casual Leave":
                row_data["opening_cl"] = opening_balance
                row_data["availed_cl"] = availed_leaves
                row_data["closing_cl"] = closing_balance
            elif leave_type == "Sick Leave":
                row_data["opening_sl"] = opening_balance
                row_data["availed_sl"] = availed_leaves
                row_data["closing_sl"] = closing_balance
            elif leave_type == "Earned Leave":
                row_data["opening_el"] = opening_balance
                row_data["availed_el"] = availed_leaves
                row_data["closing_el"] = closing_balance
            elif leave_type == "Leave Without Pay":
                row_data["opening_lop"] = 0
                row_data["availed_lop"] = abs(availed_leaves)
                row_data["closing_lop"] = 0

        leave_data.append(row_data)

    return leave_data

def get_opening_balance(employee, leave_type, from_date):
    # Get total leave ledger entries before from_date
    opening = frappe.db.sql("""
        SELECT SUM(la.leaves) as total
        FROM `tabLeave Ledger Entry` AS la
        WHERE la.employee = %s 
        AND la.leave_type = %s
        AND la.from_date < %s
        AND la.transaction_type = 'Leave Application'
    """, (employee, leave_type, from_date))

    # Get total leaves allocated for the employee and leave type
    total_leaves_allocated = frappe.db.sql("""
        SELECT SUM(total_leaves_allocated) 
        FROM `tabLeave Allocation` 
        WHERE employee = %s
        AND leave_type = %s
        AND from_date < %s
    """, (employee, leave_type, from_date))

    # Calculate opening balance as the sum of ledger entries and allocated leaves
    opening_balance = (opening[0][0] if opening and opening[0][0] is not None else 0) + \
                      (total_leaves_allocated[0][0] if total_leaves_allocated and total_leaves_allocated[0][0] is not None else 0)

    return opening_balance

def get_availed_leaves(employee, leave_type, from_date, to_date):
    availed = frappe.db.sql("""
        SELECT SUM(total_leave_days) 
        FROM `tabLeave Application` 
        WHERE employee = %s 
        AND leave_type = %s 
        AND status = 'Approved'
        AND from_date >= %s 
        AND to_date <= %s
    """, (employee, leave_type, from_date, to_date))

    return availed[0][0] if availed and availed[0][0] is not None else 0
