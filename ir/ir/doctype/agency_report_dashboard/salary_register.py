from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, getdate
from frappe import _
from frappe.utils.file_manager import get_file
from frappe.utils.background_jobs import enqueue

from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from six import BytesIO
import math

@frappe.whitelist()
def download():
    filename = 'Agency Salary Register'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name or "Sheet1"

    # Define border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Adding Header Rows
    ws.append(["INDUSTRIAS DEL RECAMBIO INDIA PRIVATE LTD"] + [""] * 31)
    ws.append(["INVOICE DETAILS FOR THE MONTH OF JANUARY- 2025"] + [""] * 31)
    ws.append(["", "", "", "", "", "", "", "", "Fixed Salary", "", "", "", "", "Earnings","", "", "", "", "", " ", "", "Employer Contribution", "", "", "", "", "", "Deduction Part", "", "", ""])
    ws.append(["S.No", "Employee", "Employee Name", "Department", "Present Days", "Paid Holiday", "Pay Days", "OT Hours", 
               "Basic", "DA", "Special Allowance", "OT Amount", "Total Gross", "Basic","Stipend", "DA", "Special Allowance", 
               "Sub Gross", "OverTime Wages", "Attendance Bonus", "Earned Gross", "PF on 13%", "ESI on 3.25%", "Insurance", 
               "Service Charge @8%", "Sub Total", "CTC", "PF on 12%", "ESI on 0.75%", "Canteen Deduction", 
               "Total Deduction", "Net Pay"])

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=32):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type="solid")  # Light gray background
            cell.border = thin_border
            
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=32):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type="solid")
            
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=32):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    # Merging Cells for Titles
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=32)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=32)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
    ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=13)
    ws.merge_cells(start_row=3, start_column=14, end_row=3, end_column=18)
    ws.merge_cells(start_row=3, start_column=19, end_row=3, end_column=21)
    ws.merge_cells(start_row=3, start_column=22, end_row=3, end_column=27)
    ws.merge_cells(start_row=3, start_column=28, end_row=3, end_column=31)

    # Fetch and Append Data
    data = get_data(args)
    for row in data:
        ws.append(row)

    # Calculate Column Totals
    totals = ["Total"] + [""] * 3 
    num_columns = len(data[0]) if data else 32  # Ensure the number of columns is correct





    for col in range(4, num_columns):  # Skip first 3 non-numeric columns
        total = sum(row[col] for row in data if isinstance(row[col], (int, float)))
        totals.append(total)

    ws.append(totals)

    total_row = ws.max_row  # Get last row index
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    # Apply borders to all cells
        # Apply formatting to the total row
    for cell in ws[total_row]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fgColor='D3D3D3', fill_type="solid")  # Yellow background
        cell.border = thin_border
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=32):
        for cell in row:
            cell.border = thin_border

    # Save to bytes
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    data = []

    filters = {
        "start_date": args.from_date,
        "end_date": args.to_date
    }

    if args.agency:
        filters["custom_agency"] = args.agency
    else:
        filters["custom_agency"] = ["!=", ""]

    salary_slips = frappe.get_all(
        "Salary Slip",
        filters=filters,
        fields=["employee", "employee_name", "department", "custom_present_days_holiday","custom_agency","custom_present_days",
                "custom_overtime", "total_working_days","payment_days","salary_structure","custom_agency_paid_holiday"], 
        order_by="employee"
    )

    i = 0
    for slip in salary_slips:
        paid_holiday = slip.custom_agency_paid_holiday
        including_holiday = frappe.db.get_value ("Agency",{"name":slip.custom_agency},["including_holiday"])
        if including_holiday == 1:
            pay = slip.custom_present_days
        else:
            pay = slip.custom_present_days_holiday
        pay_days = float(pay) + float(paid_holiday)
        basic = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_basic"])
        da = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_dearness_allowance"])
        spl = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_other_allowance"])
        ot = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_overtime_amount"])
        can_amount = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_canteen"])
        monthly_gross = frappe.db.get_value("Employee", {"name": slip.employee}, ["custom_gross"])
        ot_amount = ot *slip.custom_overtime
        gross = basic + da + spl
        ear_basic = float(pay_days) * basic
        if slip.salary_structure == "Consultant":
      
            Stipend = monthly_gross/ slip.total_working_days * slip.payment_days
        else:
            Stipend  = 0
        ear_da = pay_days * da
        ear_spl = pay_days * spl
        sub_gross = ear_basic + ear_da + ear_spl
        ot_wages = ot * slip.custom_overtime
        bonus = 300 if pay_days == slip.total_working_days else 0
        earn_gross = sub_gross + ot_wages +bonus + round(Stipend)
        pf = (ear_basic + ear_da) * 13 / 100
        esi = sub_gross * (3.25 / 100)
        insurance = 0
        service = sub_gross * (8 / 100)
        sub_total = pf + esi + insurance + service
        ctc = earn_gross + sub_total
        pf_de = math.ceil((ear_basic + ear_da) * (12 / 100))
        esi_de = math.ceil((sub_gross + ot_wages) * (0.75 / 100))
        canteen = float(pay) * can_amount
        total_de = pf_de + esi_de + canteen
        net_pay = earn_gross - total_de
        i += 1

        data.append([i, slip.employee, slip.employee_name, slip.department, 
             pay, 
             paid_holiday, 
             pay_days, 
             slip.custom_overtime, 
             round(basic, 2), 
             round(da, 2), 
             round(spl, 2), 
             round(ot_amount, 2), 
             round(gross, 2), 
             round(ear_basic, 2),
             round(Stipend), 
             round(ear_da, 2), 
             round(ear_spl, 2), 
             round(sub_gross, 2), 
             round(ot_wages, 2), 
             round(bonus, 2), 
             round(earn_gross, 2), 
             round(pf, 2), 
             round(esi, 2), 
             round(insurance, 2), 
             round(service, 2), 
             round(sub_total, 2), 
             round(ctc, 2), 
             round(pf_de, 2), 
             round(esi_de, 2), 
             round(canteen, 2), 
             round(total_de, 2), 
             round(net_pay, 2)
        ])

    return data

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
