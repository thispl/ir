# Copyright (c) 2024, TEAMPROO and contributors
# For license information, please see license.txt
# Copyright (c) 2024, Abdulla and contributors
# For license information, please see license.txt
from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


class HeadCount(Document):
	pass

@frappe.whitelist()
def download():
	filename = 'Head Count'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20 
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 20
	ws.column_dimensions['I'].width = 20
	ws.column_dimensions['J'].width = 20 
	ws.column_dimensions['K'].width = 20 
	ws.append(["SUMMARY"," "," "," "," "," "," "," "," "," "," "])
	ws.append(["S.NO","DEPARTMENT "," SHIFT 1 HEAD COUNT ","SHIFT 2 HEAD COUNT","SHIFT G HEAD COUNT","TOTAL ","SHIFT 1 OVERTIME HRS","SHIFT 2 OVERTIME HRS","SHIFT G OVERTIME HRS"," TOTAL "])
	data1= get_data(args)
	for row in data1:
		ws.append(row)
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')
	for header in ws.iter_rows(min_row=2 , max_row=2, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center

	for header in ws.iter_rows(min_row=1 , max_row=2, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(get_data(args))+2, max_row=len(get_data(args))+2, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=3, max_row=len(get_data(args))+3, min_col=1, max_col=1):
		for cell in header:
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=3, max_row=len(get_data(args))+3, min_col=3, max_col=10):
		for cell in header:
			cell.alignment = align_center

	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(get_data(args))+2, column=10).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
	ws.merge_cells(start_row=len(get_data(args))+2, start_column=1, end_row=len(get_data(args))+2, end_column=2)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_data(args):
	data = []
	row=[]

	departments = frappe.db.sql("""
		SELECT DISTINCT department
		FROM `tabAttendance`
		WHERE attendance_date BETWEEN %s AND %s
		AND docstatus != 2
	""", (args.from_date,args.to_date), as_dict=True)
	tot1=0
	tot2=0
	tot3=0
	tot4=0
	tot5=0
	tot6=0
	tot7=0
	tot8=0
	s_no = 1
	for dept in departments:
		department = dept.department
		shift1_data = frappe.db.sql("""
			SELECT COUNT(status) AS shift1_count, SUM(custom_ot_hours) AS shift1_overtime
			FROM `tabAttendance`
			WHERE attendance_date BETWEEN %s AND %s
			AND docstatus != 2
			AND in_time IS NOT NULL
			AND department = %s 
			AND shift = '1'
		""", (args.from_date,args.to_date, department), as_dict=True)

		shift1_count = shift1_data[0].shift1_count if shift1_data and shift1_data[0].shift1_count is not None else 0
		shift1_overtime = shift1_data[0].shift1_overtime if shift1_data and shift1_data[0].shift1_overtime is not None else 0.0
		shift1_overtime_hours = flt(shift1_overtime) / 10000

		shift2_data = frappe.db.sql("""
			SELECT COUNT(status) AS shift2_count, SUM(custom_ot_hours) AS shift2_overtime
			FROM `tabAttendance`
			WHERE attendance_date BETWEEN %s AND %s
			AND docstatus != 2
			AND in_time IS NOT NULL
			AND department = %s 
			AND shift = '2'
		""", (args.from_date,args.to_date, department), as_dict=True)

		shift2_count = shift2_data[0].shift2_count if shift2_data and shift2_data[0].shift2_count is not None else 0
		shift2_overtime = shift2_data[0].shift2_overtime if shift2_data and shift2_data[0].shift2_overtime is not None else 0.0
		shift2_overtime_hours = flt(shift2_overtime) / 10000
		
		shiftg_data = frappe.db.sql("""
			SELECT COUNT(status) AS shiftg_count, SUM(custom_ot_hours) AS shiftg_overtime
			FROM `tabAttendance`
			WHERE attendance_date BETWEEN %s AND %s
			AND docstatus != 2
			AND in_time IS NOT NULL
			AND department = %s 
			AND shift = 'G'
		""", (args.from_date, args.to_date, department), as_dict=True)

		shiftg_count = shiftg_data[0].shiftg_count if shiftg_data and shiftg_data[0].shiftg_count is not None else 0
		shiftg_overtime = shiftg_data[0].shiftg_overtime if shiftg_data and shiftg_data[0].shiftg_overtime is not None else 0.0
		shiftg_overtime_hours = flt(shiftg_overtime) / 10000

		total_over_time = flt(shift1_overtime_hours) + flt(shift2_overtime_hours) + flt(shiftg_overtime_hours)
		total_count = shift1_count + shift2_count + shiftg_count
		if flt(shift1_count) or flt(shift2_count) or flt(shiftg_count):
			data.append([
				s_no,
				department,
				shift1_count,
				shift2_count,
				shiftg_count,
				total_count,
				shift1_overtime_hours,
				shift2_overtime_hours,
				shiftg_overtime_hours,
				total_over_time
			])
			s_no += 1
		tot1+=shift1_count
		tot2+=shift2_count
		tot3+=shiftg_count
		tot4+=total_count
		tot5+=shift1_overtime_hours
		tot6+=shift2_overtime_hours
		tot7+=shiftg_overtime_hours
		tot8+=total_over_time

	row+=['Total','',tot1,tot2,tot3,tot4,tot5,tot6,tot7,tot8]
	data.append(row)
	return data

